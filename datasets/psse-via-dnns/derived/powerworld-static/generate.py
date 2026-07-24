"""
Generate PowerWorld-format files from PyPower case118 + a psse-via-dnns sample.

Writes a PowerWorld-compatible operating-point import set:
  buses.xlsx         — Bus sheet with sample voltage magnitudes and angles
  demand_loads.xlsx  — Load sheet with sample-scaled loads
  generators.xlsx    — Gen sheet preserving corrected PowerWorld dynamic dispatch
  equivalent_injections.xlsx
                    — Load sheet with signed static equivalents to match net injections
  ybus.xlsx          — Bus sheet, 118×118 complex Y-bus as formatted strings

Transient dynamic models and contingencies are stored inside the .pwb itself,
so no aux file is written here.

Usage:
    .venv/Scripts/python generate.py [--sample-index N] [--repo-root PATH]

--sample-index N  : which psse-via-dnns sample to use for operating-point values
                    (default 17103 = transient-compatible initial sample)
--repo-root PATH  : path to repository root (default: four levels up from this file)
"""

import argparse
import sys
from pathlib import Path

import numpy as np
import scipy.io as sio
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from pypower.case118 import case118
from pypower.ext2int import ext2int
from pypower.makeYbus import makeYbus


# ── paths ──────────────────────────────────────────────────────────────────────

HERE = Path(__file__).resolve().parent
DEFAULT_SAMPLE_INDEX = 17103


def resolve_paths(repo_root: Path) -> dict:
    psse    = repo_root / "datasets" / "psse-via-dnns" / "original" / "pypower-ieee118"
    ref     = repo_root / "datasets" / "powerworld-ieee118" / "reference"
    imports = repo_root / "datasets" / "powerworld-ieee118" / "import-files"
    imports.mkdir(parents=True, exist_ok=True)
    return {
        "psse_dir":   psse,
        "out_buses":  imports / "buses.xlsx",
        "out_loads":  imports / "demand_loads.xlsx",
        "out_gens":   imports / "generators.xlsx",
        "out_equiv":  imports / "equivalent_injections.xlsx",
        "out_ybus":   imports / "ybus.xlsx",
        "src_gens":   ref / "generators_correct.xlsx",
    }


# ── Y-bus ───────────────────────────────────────────────────────────────────────

def make_ybus(case: dict) -> np.ndarray:
    """Build full complex admittance matrix via pypower.makeYbus (handles taps, shunts, phase shifters)."""
    case_i = ext2int(case)
    Y, _, _ = makeYbus(case_i["baseMVA"], case_i["bus"], case_i["branch"])
    return Y.toarray()


def fmt_complex(z: complex) -> str:
    """Format Y-bus entry as PowerWorld string: 'A.4f ± jB.4f'."""
    a = z.real
    b = z.imag
    if b < 0:
        return f"{a:.4f} - j{abs(b):.4f}"
    else:
        return f"{a:.4f} + j{b:.4f}"


# ── load scale ─────────────────────────────────────────────────────────────────

def derive_load_scale(case: dict, Pi_pu: np.ndarray) -> float:
    """
    Compute load scale factor for one sample using pure load buses (no generator).

    Pi_pu : 1-D array of net bus active power injection (pu) for the sample,
            length = n_buses.  Negative values mean net load.
    Returns: scale factor s such that P_load_actual = s * P_load_base.
    """
    gen_buses = set(case["gen"][:, 0].astype(int))
    pure_load_idx = [
        i for i in range(case["bus"].shape[0])
        if case["bus"][i, 2] > 0 and int(case["bus"][i, 0]) not in gen_buses
    ]
    if not pure_load_idx:
        raise ValueError("no pure load buses found; cannot derive load scale")
    pd_base_pu = case["bus"][pure_load_idx, 2] / case["baseMVA"]
    pi_sample  = Pi_pu[pure_load_idx]
    scale = float(np.mean(-pi_sample / pd_base_pu))
    return scale


# ── demand loads ───────────────────────────────────────────────────────────────

_LOAD_COLS = [
    "Number of Bus", "Name of Bus", "Area Name of Load", "Zone Name of Load",
    "ID", "Status", "MW", "Mvar", "MVA", "S MW", "S Mvar",
    "Dist Status", "Dist MW Input", "Dist Mvar Input",
]


def build_load_rows(
    case: dict,
    scale: float,
) -> tuple[list[list], dict[int, tuple[float, float]], dict[int, tuple[float, float]]]:
    """Build rows for load records that already exist in the PowerWorld PWB."""
    gen_buses = set(case["gen"][:, 0].astype(int))
    rows = []
    load_by_bus = {}
    omitted_by_bus = {}
    for row in case["bus"]:
        bus_num = int(row[0])
        pd_mw = float(row[2])
        qd_mvar = float(row[3])
        if pd_mw <= 0:
            continue
        # PowerWorld will not create missing Load records while pasting in Run
        # Mode. These generator-bus, QD=0 loads are absent from the source PWB,
        # so leave them out and preserve their net injection through Gen MW.
        if bus_num in gen_buses and abs(qd_mvar) < 1e-9:
            omitted_by_bus[bus_num] = (pd_mw * scale, qd_mvar * scale)
            continue
        mw_scaled = pd_mw * scale
        mvar_scaled = qd_mvar * scale
        load_by_bus[bus_num] = (mw_scaled, mvar_scaled)
        rows.append([
            bus_num,
            str(bus_num),   # Name of Bus: matches original (bare number, not "Bus N")
            "Main",         # Area Name of Load: single area in original
            "1",            # Zone Name of Load: single zone in original
            "1",            # ID
            "Closed",       # Status
            round(mw_scaled,   4),
            round(mvar_scaled, 4),
            "",             # MVA (blank in original)
            "",             # S MW
            "",             # S Mvar
            "Closed",       # Dist Status
            "",             # Dist MW Input
            "",             # Dist Mvar Input
        ])
    return rows, load_by_bus, omitted_by_bus


def write_demand_loads(path: Path, rows: list[list]) -> None:
    """Write demand_loads.xlsx replicating PowerWorld Load sheet format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Load"

    # Row 1: section label
    ws.cell(1, 1, "Load").font = Font(bold=True)
    # Row 2: column headers
    for ci, col in enumerate(_LOAD_COLS, start=1):
        ws.cell(2, ci, col)
    # Data rows starting at row 3
    for ri, row in enumerate(rows, start=3):
        for ci, val in enumerate(row, start=1):
            ws.cell(ri, ci, val)

    wb.save(path)
    print(f"  Wrote {len(rows)} load rows: {path}")


# ── generators ─────────────────────────────────────────────────────────────────

_GEN_COLS = [
    "Number of Bus", "Name of Bus", "ID", "Status",
    "Gen MW", "Gen Mvar", "Min MW", "Max MW",
    "AGC", "AVR", "RegBus Num", "Set Volt",
    "Min Mvar", "Max Mvar",
    "Enforce MW Limits", "Part. Factor", "Cost Model",
]


def read_reference_generator_rows(path: Path, case: dict, Vm_pu: np.ndarray) -> list[list]:
    """Read corrected generator rows, preserving dispatch but matching sample voltage targets."""
    if not path.exists():
        raise FileNotFoundError(f"reference generator workbook not found: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[2]]
    if headers[:len(_GEN_COLS)] != _GEN_COLS:
        raise ValueError(f"unexpected generator workbook columns in {path}")
    rows = []
    bus_ids = case["bus"][:, 0].astype(int)
    sample_vm_by_bus = {int(bus_num): float(Vm_pu[i]) for i, bus_num in enumerate(bus_ids)}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        out = list(row[:len(_GEN_COLS)])
        bus_num = int(out[0])
        if bus_num in sample_vm_by_bus:
            out[11] = round(sample_vm_by_bus[bus_num], 4)
        rows.append(out)
    return rows


def write_generators(path: Path, rows: list[list]) -> None:
    """Write generators.xlsx from corrected PowerWorld dynamic generator rows."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Gen"

    ws.cell(1, 1, "Gen").font = Font(bold=True)
    for ci, col in enumerate(_GEN_COLS, start=1):
        ws.cell(2, ci, col)
    for ri, row in enumerate(rows, start=3):
        for ci, val in enumerate(row, start=1):
            ws.cell(ri, ci, val)

    wb.save(path)
    print(f"  Wrote {len(rows)} generator rows: {path}")


# ── equivalent injections ──────────────────────────────────────────────────────

_EQUIV_LOAD_ID = "E"


def build_equivalent_injection_rows(
    case: dict,
    Pi_pu: np.ndarray,
    Qi_pu: np.ndarray,
    load_by_bus: dict[int, tuple[float, float]],
    gen_rows: list[list],
) -> tuple[list[list], dict[int, tuple[float, float]]]:
    """
    Build signed Load rows that make bus net injection match the sample.

    PowerWorld load sign convention is positive consumption, so an equivalent
    load of -10 MW represents a +10 MW static injection at that bus.
    """
    baseMVA = float(case["baseMVA"])
    bus_nums = case["bus"][:, 0].astype(int)

    gen_by_bus: dict[int, tuple[float, float]] = {}
    for row in gen_rows:
        bus_num = int(row[0])
        gen_mw, gen_mvar = gen_by_bus.get(bus_num, (0.0, 0.0))
        gen_by_bus[bus_num] = (gen_mw + float(row[4]), gen_mvar + float(row[5]))

    rows = []
    equiv_by_bus = {}
    for i, bus_num in enumerate(bus_nums):
        target_net_mw = float(Pi_pu[i]) * baseMVA
        target_net_mvar = float(Qi_pu[i]) * baseMVA
        gen_mw, gen_mvar = gen_by_bus.get(int(bus_num), (0.0, 0.0))
        load_mw, load_mvar = load_by_bus.get(int(bus_num), (0.0, 0.0))

        equiv_mw = gen_mw - load_mw - target_net_mw
        equiv_mvar = gen_mvar - load_mvar - target_net_mvar
        if abs(equiv_mw) < 5e-5 and abs(equiv_mvar) < 5e-5:
            continue
        equiv_by_bus[int(bus_num)] = (equiv_mw, equiv_mvar)
        rows.append([
            int(bus_num),
            str(int(bus_num)),
            "Main",
            "1",
            _EQUIV_LOAD_ID,
            "Closed",
            round(equiv_mw, 4),
            round(equiv_mvar, 4),
            "",
            round(equiv_mw, 4),
            round(equiv_mvar, 4),
            "Closed",
            "",
            "",
        ])
    return rows, equiv_by_bus


def write_equivalent_injections(path: Path, rows: list[list]) -> None:
    """Write static equivalent signed loads used to match sample net injections."""
    write_demand_loads(path, rows)


# ── bus workbook ───────────────────────────────────────────────────────────────

_BUS_COLS = [
    "Number", "Name", "Area Name", "Nom kV", "PU Volt", "Volt (kV)",
    "Angle (Deg)", "Load MW", "Load Mvar", "Gen MW", "Gen Mvar",
    "Switched Shunts Mvar", "Act G Shunt MW", "Act B Shunt Mvar",
    "Area Num", "Zone Num",
]


def _blank_if_zero(value: float) -> float | None:
    return None if abs(value) < 1e-9 else round(value, 4)


def write_buses(
    path: Path,
    case: dict,
    Vm_pu: np.ndarray,
    Va_rad: np.ndarray,
    load_by_bus: dict[int, tuple[float, float]],
    equiv_by_bus: dict[int, tuple[float, float]],
    gen_rows: list[list],
) -> None:
    """Write a PowerWorld Bus sheet carrying the selected sample V/theta state."""
    gen_by_bus = {}
    for row in gen_rows:
        bus_num = int(row[0])
        gen_mw, gen_mvar = gen_by_bus.get(bus_num, (0.0, 0.0))
        gen_by_bus[bus_num] = (gen_mw + float(row[4]), gen_mvar + float(row[5]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Bus"

    ws.cell(1, 1, "Bus").font = Font(bold=True)
    for ci, col in enumerate(_BUS_COLS, start=1):
        ws.cell(2, ci, col)

    for i, row in enumerate(case["bus"]):
        bus_num = int(row[0])
        nom_kv = float(row[9])
        vm = float(Vm_pu[i])
        va_deg = float(np.degrees(Va_rad[i]))
        load_mw, load_mvar = load_by_bus.get(bus_num, (0.0, 0.0))
        equiv_mw, equiv_mvar = equiv_by_bus.get(bus_num, (0.0, 0.0))
        gen_mw, gen_mvar = gen_by_bus.get(bus_num, (0.0, 0.0))
        out = [
            bus_num,
            str(bus_num),
            "Main",
            _blank_if_zero(nom_kv),
            round(vm, 6),
            round(vm * nom_kv, 4),
            round(va_deg, 6),
            _blank_if_zero(load_mw + equiv_mw),
            _blank_if_zero(load_mvar + equiv_mvar),
            _blank_if_zero(gen_mw),
            _blank_if_zero(gen_mvar),
            None,
            _blank_if_zero(float(row[4])),
            _blank_if_zero(float(row[5])),
            int(row[6]),
            int(row[10]),
        ]
        for ci, val in enumerate(out, start=1):
            ws.cell(i + 3, ci, val)

    wb.save(path)
    print(f"  Wrote {case['bus'].shape[0]} bus rows: {path}")


# ── Y-bus workbook ──────────────────────────────────────────────────────────────

def write_ybus(path: Path, case: dict, Y: np.ndarray) -> None:
    """Write ybus.xlsx replicating PowerWorld Bus sheet format.

    Layout: 2 header rows, then one row per bus.
    Columns: [Bus #, Bus Name, Y11, Y12, ..., Y1_118].
    """
    bus     = case["bus"]
    n_buses = bus.shape[0]
    bus_ids = bus[:, 0].astype(int)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bus"

    # Row 1: section label
    ws.cell(1, 1, "Bus").font = Font(bold=True)
    # Row 2: column headers
    ws.cell(2, 1, "Bus #")
    ws.cell(2, 2, "Bus Name")
    for ci, bid in enumerate(bus_ids, start=3):
        ws.cell(2, ci, bid)

    # Data rows
    for i in range(n_buses):
        ri = i + 3
        ws.cell(ri, 1, int(bus_ids[i]))
        ws.cell(ri, 2, f"Bus {bus_ids[i]}")
        for j in range(n_buses):
            z = Y[i, j]
            if z == 0:
                continue
            ws.cell(ri, j + 3, fmt_complex(z))

    wb.save(path)
    print(f"  Wrote {n_buses}x{n_buses} Y-bus: {path}")


def print_validation_summary(
    sample_index: int,
    scale: float,
    case: dict,
    Pi_pu: np.ndarray,
    Qi_pu: np.ndarray,
    Vm_pu: np.ndarray,
    Va_rad: np.ndarray,
    load_by_bus: dict[int, tuple[float, float]],
    omitted_load_by_bus: dict[int, tuple[float, float]],
    equiv_by_bus: dict[int, tuple[float, float]],
    gen_rows: list[list],
) -> None:
    total_load_mw = sum(mw for mw, _ in load_by_bus.values())
    total_load_mvar = sum(mvar for _, mvar in load_by_bus.values())
    omitted_load_mw = sum(mw for mw, _ in omitted_load_by_bus.values())
    omitted_load_mvar = sum(mvar for _, mvar in omitted_load_by_bus.values())
    equiv_abs_mw = sum(abs(mw) for mw, _ in equiv_by_bus.values())
    equiv_abs_mvar = sum(abs(mvar) for _, mvar in equiv_by_bus.values())
    gen_by_bus = {int(row[0]): row for row in gen_rows}
    bus69 = gen_by_bus.get(69)
    negative = [row for row in gen_rows if float(row[4]) < -1e-9]
    baseMVA = float(case["baseMVA"])
    bus_ids = case["bus"][:, 0].astype(int)
    residual_p = []
    residual_q = []
    gen_totals: dict[int, tuple[float, float]] = {}
    for row in gen_rows:
        bus_num = int(row[0])
        gen_mw, gen_mvar = gen_totals.get(bus_num, (0.0, 0.0))
        gen_totals[bus_num] = (gen_mw + float(row[4]), gen_mvar + float(row[5]))
    for i, bus_num_raw in enumerate(bus_ids):
        bus_num = int(bus_num_raw)
        gen_mw, gen_mvar = gen_totals.get(bus_num, (0.0, 0.0))
        load_mw, load_mvar = load_by_bus.get(bus_num, (0.0, 0.0))
        equiv_mw, equiv_mvar = equiv_by_bus.get(bus_num, (0.0, 0.0))
        net_mw = gen_mw - load_mw - equiv_mw
        net_mvar = gen_mvar - load_mvar - equiv_mvar
        residual_p.append(abs(net_mw / baseMVA - float(Pi_pu[i])))
        residual_q.append(abs(net_mvar / baseMVA - float(Qi_pu[i])))

    print("Validation summary:")
    print(f"  Sample index: {sample_index}")
    print(f"  Load scale factor: {scale:.6f}")
    print(f"  Total represented load: {total_load_mw:.3f} MW, {total_load_mvar:.3f} Mvar")
    if omitted_load_by_bus:
        buses = ", ".join(str(bus_num) for bus_num in sorted(omitted_load_by_bus))
        print(
            "  Omitted missing-PWB load records: "
            f"{len(omitted_load_by_bus)} buses ({omitted_load_mw:.3f} MW, "
            f"{omitted_load_mvar:.3f} Mvar): {buses}"
        )
    if bus69 is None:
        print("  WARNING: Bus 69 generator row was not written")
    else:
        print(f"  Bus 69 generator: {float(bus69[4]):.4f} MW, {float(bus69[5]):.4f} Mvar")
    print(
        "  Static equivalent injections: "
        f"{len(equiv_by_bus)} buses, |MW| sum {equiv_abs_mw:.3f}, "
        f"|Mvar| sum {equiv_abs_mvar:.3f}"
    )
    print(
        "  Net injection reconstruction residual: "
        f"P <= {max(residual_p):.3e} pu, Q <= {max(residual_q):.3e} pu"
    )
    print(
        "  Sample voltage range: "
        f"{float(np.min(Vm_pu)):.6f} to {float(np.max(Vm_pu)):.6f} pu"
    )
    print(
        "  Sample angle range: "
        f"{float(np.degrees(np.min(Va_rad))):.6f} to "
        f"{float(np.degrees(np.max(Va_rad))):.6f} deg"
    )
    print(f"  Negative-MW generators: {len(negative)}")
    if negative:
        details = ", ".join(f"{int(row[0])}#{row[2]}={float(row[4]):.4f} MW" for row in negative)
        print(f"  WARNING: {details}")


# ── main ───────────────────────────────────────────────────────────────────────

def run(repo_root: Path, sample_index: int) -> int:
    paths = resolve_paths(repo_root)
    psse_dir = paths["psse_dir"]

    if not psse_dir.exists():
        print(f"ERROR: psse_dir not found: {psse_dir}", file=sys.stderr)
        return 1
    inj_path = psse_dir / "Pi_Qi_injections.mat"
    voltage_path = psse_dir / "V_magnitudes_angles.mat"
    if not inj_path.exists():
        print(f"ERROR: not found: {inj_path}", file=sys.stderr)
        return 1
    if not voltage_path.exists():
        print(f"ERROR: not found: {voltage_path}", file=sys.stderr)
        return 1
    if not paths["src_gens"].exists():
        print(f"ERROR: not found: {paths['src_gens']}", file=sys.stderr)
        return 1

    case    = case118()
    n_buses = case["bus"].shape[0]

    inj = sio.loadmat(str(inj_path))["data"]
    n_samples = inj.shape[0]
    if not (0 <= sample_index < n_samples):
        print(f"ERROR: --sample-index {sample_index} out of range [0, {n_samples - 1}]",
              file=sys.stderr)
        return 1
    print(f"Loading sample {sample_index} from psse-via-dnns ...")
    Pi_pu = inj[sample_index, :n_buses]
    Qi_pu = inj[sample_index, n_buses:]
    voltage = sio.loadmat(str(voltage_path))["data"]
    if voltage.shape[0] != n_samples or voltage.shape[1] != 2 * n_buses:
        print(
            f"ERROR: unexpected V_magnitudes_angles.mat shape {voltage.shape}; "
            f"expected ({n_samples}, {2 * n_buses})",
            file=sys.stderr,
        )
        return 1
    Vm_pu = voltage[sample_index, :n_buses]
    Va_rad = voltage[sample_index, n_buses:]

    scale = derive_load_scale(case, Pi_pu)
    print(f"  Load scale factor: {scale:.4f}")

    HERE.mkdir(parents=True, exist_ok=True)

    print("Building Y-bus ...")
    Y = make_ybus(case)

    load_rows, load_by_bus, omitted_load_by_bus = build_load_rows(case, scale)

    print("Writing demand_loads.xlsx ...")
    write_demand_loads(paths["out_loads"], load_rows)

    print("Writing generators.xlsx ...")
    gen_rows = read_reference_generator_rows(paths["src_gens"], case, Vm_pu)
    write_generators(paths["out_gens"], gen_rows)

    print("Writing equivalent_injections.xlsx ...")
    equiv_rows, equiv_by_bus = build_equivalent_injection_rows(case, Pi_pu, Qi_pu, load_by_bus, gen_rows)
    write_equivalent_injections(paths["out_equiv"], equiv_rows)

    print("Writing buses.xlsx ...")
    write_buses(paths["out_buses"], case, Vm_pu, Va_rad, load_by_bus, equiv_by_bus, gen_rows)

    print("Writing ybus.xlsx ...")
    write_ybus(paths["out_ybus"], case, Y)

    print_validation_summary(
        sample_index,
        scale,
        case,
        Pi_pu,
        Qi_pu,
        Vm_pu,
        Va_rad,
        load_by_bus,
        omitted_load_by_bus,
        equiv_by_bus,
        gen_rows,
    )

    print("Done.")
    return 0


if __name__ == "__main__":
    default_repo = Path(__file__).resolve().parents[4]
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--sample-index", type=int, default=DEFAULT_SAMPLE_INDEX,
        help=f"Index of psse-via-dnns sample to use (default: {DEFAULT_SAMPLE_INDEX})",
    )
    parser.add_argument(
        "--repo-root", type=Path, default=default_repo,
        help=f"Repository root (default: {default_repo})",
    )
    args = parser.parse_args()
    sys.exit(run(args.repo_root, args.sample_index))
