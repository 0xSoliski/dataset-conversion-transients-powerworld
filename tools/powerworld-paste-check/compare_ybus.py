"""
Compare a PowerWorld-exported Ybus workbook against the generated PYPOWER Ybus.

Run after exporting the Bus Admittance Matrix from PowerWorld:

    .venv/Scripts/python tools/powerworld-paste-check/compare_ybus.py \
        --powerworld-ybus <powerworld-ybus-export.xlsx>
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
TARGET_YBUS = REPO_ROOT / "datasets" / "powerworld-ieee118" / "import-files" / "ybus.xlsx"


def parse_complex(value: object) -> complex:
    if value is None or value == "":
        return 0j
    if isinstance(value, (int, float)):
        return complex(value, 0.0)
    compact = str(value).strip().replace(" ", "").replace("j", "")
    match = re.match(r"^([+-]?\d+(?:\.\d+)?)([+-]\d+(?:\.\d+)?)$", compact)
    if not match:
        raise ValueError(f"cannot parse complex Ybus cell: {value!r}")
    return complex(float(match.group(1)), float(match.group(2)))


def read_ybus(path: Path) -> tuple[list[int], np.ndarray]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    size = ws.max_row - 2
    bus_ids: list[int] = []
    ybus = np.zeros((size, size), dtype=complex)

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        if row[0] is None:
            continue
        bus_ids.append(int(row[0]))
        for col_idx, value in enumerate(row[2 : 2 + size]):
            ybus[row_idx, col_idx] = parse_complex(value)

    return bus_ids, ybus


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--powerworld-ybus", type=Path, required=True)
    parser.add_argument("--target-ybus", type=Path, default=TARGET_YBUS)
    parser.add_argument("--tolerance", type=float, default=1e-3)
    args = parser.parse_args()

    target_ids, target = read_ybus(args.target_ybus)
    actual_ids, actual = read_ybus(args.powerworld_ybus)
    if target_ids != actual_ids:
        print("Bus order differs between Ybus files.")
        print(f"target first/last: {target_ids[:5]} ... {target_ids[-5:]}")
        print(f"actual first/last: {actual_ids[:5]} ... {actual_ids[-5:]}")
        return 1

    diff = actual - target
    abs_diff = np.abs(diff)
    max_idx = np.unravel_index(np.argmax(abs_diff), abs_diff.shape)
    count = int(np.sum(abs_diff > args.tolerance))

    print(f"Compared {len(target_ids)}x{len(target_ids)} Ybus matrices")
    print(
        "Max |diff|: "
        f"{abs_diff[max_idx]:.6g} at bus {target_ids[max_idx[0]]}-{target_ids[max_idx[1]]}"
    )
    print(f"Entries above {args.tolerance:g}: {count}")

    if count:
        print("Largest differences:")
        entries = [
            (abs_diff[i, j], i, j)
            for i in range(abs_diff.shape[0])
            for j in range(abs_diff.shape[1])
            if abs_diff[i, j] > args.tolerance
        ]
        for _, i, j in sorted(entries, reverse=True)[:30]:
            print(
                f"  {target_ids[i]}-{target_ids[j]}: "
                f"target={target[i, j]:.6g}, actual={actual[i, j]:.6g}, "
                f"diff={diff[i, j]:.6g}"
            )
        return 1

    print("Result: PowerWorld Ybus matches generated PYPOWER Ybus within tolerance.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
