"""
Compare PowerWorld case-information exports against the generated paste targets.

Export Bus, Load, and Gen case-information displays from PowerWorld after
pasting the generated files but before running transient stability. Then run:

    .venv/Scripts/python tools/powerworld-paste-check/compare_pasted_state.py \
        --bus-export <powerworld-bus-export.xlsx> \
        --load-export <powerworld-load-export.xlsx> \
        --gen-export <powerworld-gen-export.xlsx>
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
TARGET_DIR = REPO_ROOT / "datasets" / "powerworld-ieee118" / "import-files"


def read_table(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[2]]
    rows: list[dict[str, object]] = []
    for values in ws.iter_rows(min_row=3, values_only=True):
        if values[0] is None:
            continue
        rows.append({headers[i]: values[i] for i in range(min(len(headers), len(values)))})
    return headers, rows


def number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_key(row: dict[str, object]) -> tuple[int, str]:
    return int(number(row["Number of Bus"])), text(row["ID"])


def gen_key(row: dict[str, object]) -> tuple[int, str]:
    return int(number(row["Number of Bus"])), text(row["ID"])


def bus_key(row: dict[str, object]) -> int:
    return int(number(row["Number"]))


def top_diffs(
    label: str,
    expected: dict[object, dict[str, object]],
    actual: dict[object, dict[str, object]],
    fields: Iterable[str],
    tolerance: float,
) -> int:
    problems = []
    missing = sorted(set(expected) - set(actual))
    extra = sorted(set(actual) - set(expected))
    for key in sorted(set(expected) & set(actual)):
        for field in fields:
            diff = number(actual[key].get(field)) - number(expected[key].get(field))
            if abs(diff) > tolerance:
                problems.append((abs(diff), key, field, number(expected[key].get(field)), number(actual[key].get(field)), diff))

    print(f"{label}:")
    print(f"  expected rows: {len(expected)}, actual rows: {len(actual)}")
    if missing:
        print(f"  missing rows: {missing[:20]}" + (" ..." if len(missing) > 20 else ""))
    if extra:
        print(f"  extra rows: {extra[:20]}" + (" ..." if len(extra) > 20 else ""))
    if problems:
        print("  largest numeric differences:")
        for _, key, field, exp, act, diff in sorted(problems, reverse=True)[:20]:
            print(f"    {key} {field}: expected {exp:.6g}, actual {act:.6g}, diff {diff:.6g}")
    else:
        print(f"  numeric fields match within {tolerance:g}")
    print()
    return len(missing) + len(problems)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--bus-export", type=Path, required=True)
    parser.add_argument("--load-export", type=Path, required=True)
    parser.add_argument("--gen-export", type=Path, required=True)
    parser.add_argument("--tolerance", type=float, default=1e-3)
    args = parser.parse_args()

    _, target_bus_rows = read_table(TARGET_DIR / "buses.xlsx")
    _, target_demand_rows = read_table(TARGET_DIR / "demand_loads.xlsx")
    _, target_equiv_rows = read_table(TARGET_DIR / "equivalent_injections.xlsx")
    _, target_gen_rows = read_table(TARGET_DIR / "generators.xlsx")

    _, actual_bus_rows = read_table(args.bus_export)
    _, actual_load_rows = read_table(args.load_export)
    _, actual_gen_rows = read_table(args.gen_export)

    expected_loads = {load_key(row): row for row in [*target_demand_rows, *target_equiv_rows]}
    actual_loads = {
        load_key(row): row
        for row in actual_load_rows
        if load_key(row) in expected_loads
    }

    failures = 0
    failures += top_diffs(
        "Bus paste state",
        {bus_key(row): row for row in target_bus_rows},
        {bus_key(row): row for row in actual_bus_rows},
        ["PU Volt", "Angle (Deg)", "Load MW", "Load Mvar", "Gen MW", "Gen Mvar"],
        args.tolerance,
    )
    failures += top_diffs(
        "Load paste state",
        expected_loads,
        actual_loads,
        ["MW", "Mvar", "S MW", "S Mvar"],
        args.tolerance,
    )
    failures += top_diffs(
        "Gen paste state",
        {gen_key(row): row for row in target_gen_rows},
        {gen_key(row): row for row in actual_gen_rows},
        ["Gen MW", "Gen Mvar", "Set Volt"],
        args.tolerance,
    )

    if failures:
        print("Result: paste/export state does not match generated targets.")
        return 1
    print("Result: paste/export state matches generated targets. Solver mismatch is model/control mismatch, not paste failure.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
