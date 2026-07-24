"""Convert a PowerWorld transient-stability export into model-ready datasets."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterable

import numpy as np
import scipy.io
from openpyxl import load_workbook

SHEET_NAME = "TSTimePointResult"
DEFAULT_BUS_COUNT = 118
BLOCK_LABELS = (
    "V pu",
    "Gen MW",
    "Gen Mvar",
    "Load MW",
    "Load Mvar",
    "V Angle (rad)",
)


def _expected_headers(bus_count: int) -> list[str]:
    return [
        f"Bus {bus_number} {label}"
        for label in BLOCK_LABELS
        for bus_number in range(1, bus_count + 1)
    ]


def read_powerworld_export(
    input_path: Path | str,
    *,
    bus_count: int = DEFAULT_BUS_COUNT,
    skip_data_rows: int = 2,
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Read and validate the PowerWorld ``TSTimePointResult`` worksheet.

    ``skip_data_rows=2`` preserves the historical MATLAB converter contract,
    which discards the first two numeric time points after importing the sheet.
    """
    input_path = Path(input_path)
    if not input_path.is_file():
        raise FileNotFoundError(f"PowerWorld export not found: {input_path}")
    if bus_count <= 0:
        raise ValueError("bus_count must be positive")
    if skip_data_rows < 0:
        raise ValueError("skip_data_rows must be non-negative")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"{input_path} does not contain the required '{SHEET_NAME}' worksheet."
            )

        worksheet = workbook[SHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)
        try:
            title_row = next(rows)
            header_row = next(rows)
        except StopIteration as exc:
            raise ValueError(f"{input_path} must contain two header rows.") from exc

        if not title_row or title_row[0] != SHEET_NAME:
            raise ValueError(
                f"{input_path} has an invalid title row; expected '{SHEET_NAME}'."
            )

        expected_headers = ["Time", *_expected_headers(bus_count)]
        actual_headers = list(header_row[: len(expected_headers)])
        if actual_headers != expected_headers:
            mismatch = next(
                (
                    index
                    for index, (actual, expected) in enumerate(
                        zip(actual_headers, expected_headers, strict=False)
                    )
                    if actual != expected
                ),
                min(len(actual_headers), len(expected_headers)),
            )
            expected = expected_headers[mismatch] if mismatch < len(expected_headers) else "<end>"
            actual = actual_headers[mismatch] if mismatch < len(actual_headers) else "<missing>"
            raise ValueError(
                f"{input_path} column {mismatch + 1} is {actual!r}; expected {expected!r}."
            )

        expected_width = 1 + len(BLOCK_LABELS) * bus_count
        values: list[tuple[object, ...]] = []
        for row_number, row in enumerate(rows, start=3):
            materialized = tuple(row[:expected_width])
            if not any(value is not None for value in materialized):
                continue
            if len(materialized) != expected_width or any(
                value is None for value in materialized
            ):
                raise ValueError(
                    f"{input_path} row {row_number} must contain {expected_width} numeric values."
                )
            values.append(materialized)
    finally:
        workbook.close()

    if len(values) <= skip_data_rows:
        raise ValueError(f"{input_path} contains no transient samples.")

    try:
        matrix = np.asarray(values[skip_data_rows:], dtype=np.float64)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{input_path} contains non-numeric transient data.") from exc
    if not np.isfinite(matrix).all():
        raise ValueError(f"{input_path} contains NaN or infinite values.")

    time = matrix[:, 0]
    blocks = matrix[:, 1:].reshape(matrix.shape[0], len(BLOCK_LABELS), bus_count)
    voltage, generation_p, generation_q, load_p, load_q, angle = (
        blocks[:, index, :] for index in range(len(BLOCK_LABELS))
    )

    features = np.hstack(((generation_p - load_p) / 100.0, (generation_q - load_q) / 100.0))
    labels = np.hstack((voltage, angle))
    return time, features, labels


def write_dataset(
    output_basename: Path | str,
    *,
    features: np.ndarray,
    labels: np.ndarray,
) -> tuple[Path, Path]:
    """Write MATLAB and CSV forms using the shared ``features``/``labels`` contract."""
    output_basename = Path(output_basename)
    output_basename.parent.mkdir(parents=True, exist_ok=True)
    mat_path = output_basename.with_suffix(".mat")
    csv_path = output_basename.with_suffix(".csv")

    scipy.io.savemat(mat_path, {"features": features, "labels": labels})
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerows(np.hstack((features, labels)))
    return mat_path, csv_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert a PowerWorld TSTimePointResult workbook to MAT and CSV datasets."
    )
    parser.add_argument("input_xlsx", type=Path)
    parser.add_argument("output_basename", type=Path)
    parser.add_argument("--bus-count", type=int, default=DEFAULT_BUS_COUNT)
    parser.add_argument(
        "--skip-data-rows",
        type=int,
        default=2,
        help="Initial numeric time points to discard (default: 2 for MATLAB compatibility).",
    )
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    _, features, labels = read_powerworld_export(
        args.input_xlsx,
        bus_count=args.bus_count,
        skip_data_rows=args.skip_data_rows,
    )
    mat_path, csv_path = write_dataset(
        args.output_basename,
        features=features,
        labels=labels,
    )
    print(f"Wrote {features.shape[0]} samples to {mat_path} and {csv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
